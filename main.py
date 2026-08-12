# -*- coding: utf-8 -*-
"""
Automatizacion de carga de fotos - Refacturaciones Aguas Andinas
==================================================================

Que hace:
    1. Lee las fotos de una carpeta de mes (ej. "Junio 2026").
    2. Agrupa las fotos por ID de cliente (el numero antes de sufijos
       tipo (1), (2), -1, _1, " 1", etc).
    3. Para cada ID: busca el cliente en el sistema, revisa si YA tiene
       imagenes adjuntas. Si tiene imagenes, cierra UNICAMENTE la ventana
       emergente de imagenes sin afectar la ventana principal y pasa al siguiente ID.
       Si no tiene imagenes, adjunta las fotos nuevas (maximo 3, maximo 5MB cada una)
       forzando el foco en la casilla 'Nombre:' (Alt+N) y limpiando el campo.
    4. Deja un registro (procesados.csv) y genera un reporte Excel (.xlsx)
       con el ID, si se subio, si necesita revision y el detalle.

Requiere:
    pip install pywinauto openpyxl pywin32
"""

import csv
import ctypes
import os
import re
import sys
import time
import traceback
from dataclasses import dataclass, field
from datetime import datetime

# En Windows con escalado de pantalla (>100%), declarar DPI awareness antes de pywinauto
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(2)
except Exception:
    try:
        ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass

# ---------------------------------------------------------------------------
# CONFIGURACIÓN GENERAL
# ---------------------------------------------------------------------------

CARPETA_FOTOS_BASE = r"C:\Users\eramirez\Documents\FOTOS"

MAX_ARCHIVOS_POR_ID = 3
MAX_TAMANO_MB = 5

ARCHIVO_REGISTRO = "procesados.csv"

# Textos / títulos de la app (para pywinauto)
TITULO_VENTANA_PRINCIPAL_REGEX = r".*Consultas.*"
NOMBRE_CAMPO_NRO = "Nro:"
NOMBRE_BOTON_ADJUNTAR = "Adjuntar imágen"
NOMBRE_BOTON_VER_IMAGENES = "Ver imágen(es)"
TITULO_VENTANA_ADJUNTAR = "Adjuntar imágen"
TITULO_VENTANA_IMAGENES_REGEX = r".*Imá?genes del cliente.*"
NOMBRE_BOTON_EXPLORAR = "..."
NOMBRE_BOTON_AGREGAR = "Agregar"
NOMBRE_BOTON_CERRAR_POPUP = "Cerrar"
NOMBRE_BOTON_ACEPTAR_POPUP = "Aceptar"

TIMEOUT_ESPERA = 15
PAUSA_CORTA = 0.4

# ---------------------------------------------------------------------------
# ESTRUCTURAS DE DATOS
# ---------------------------------------------------------------------------


@dataclass
class GrupoID:
    id_cliente: str
    archivos_validos: list = field(default_factory=list)
    archivos_excedente: list = field(default_factory=list)
    archivos_muy_grandes: list = field(default_factory=list)


@dataclass
class ResultadoID:
    id_cliente: str
    estado: str          # OK / OMITIDO_YA_TENIA / NO_ENCONTRADO / ERROR / EXCEDENTE / MUY_GRANDE
    detalle: str = ""
    archivos: list = field(default_factory=list)


# ---------------------------------------------------------------------------
# PASO 1: ESCANEO Y AGRUPAMIENTO DE ARCHIVOS
# ---------------------------------------------------------------------------

PATRON_SUFIJO = re.compile(
    r"""^
    (?P<id>\d+)                      # el ID: uno o mas digitos al inicio
    (?:                               # sufijo opcional:
        \s*\(\d+\)                    #   (1)  (2) ...
        |\s*-\s*\d+                   #   -1   - 1 ...
        |\s*_\s*\d+                   #   _1
    )?
    $""",
    re.VERBOSE,
)

EXTENSIONES_IMAGEN = {".jpg", ".jpeg", ".png", ".bmp"}


def extraer_id_limpio(nombre_archivo_sin_ext: str) -> str | None:
    m = PATRON_SUFIJO.match(nombre_archivo_sin_ext.strip())
    if not m:
        return None
    return m.group("id")


def escanear_carpeta(carpeta_mes: str):
    grupos: dict[str, GrupoID] = {}
    no_reconocidos: list[str] = []

    if not os.path.isdir(carpeta_mes):
        raise FileNotFoundError(f"No existe la carpeta: {carpeta_mes}")

    for nombre in sorted(os.listdir(carpeta_mes)):
        ruta_completa = os.path.join(carpeta_mes, nombre)
        if not os.path.isfile(ruta_completa):
            continue

        base, ext = os.path.splitext(nombre)
        if ext.lower() not in EXTENSIONES_IMAGEN:
            continue

        id_limpio = extraer_id_limpio(base)
        if id_limpio is None:
            no_reconocidos.append(nombre)
            continue

        grupo = grupos.setdefault(id_limpio, GrupoID(id_cliente=id_limpio))

        tamano_mb = os.path.getsize(ruta_completa) / (1024 * 1024)
        if tamano_mb > MAX_TAMANO_MB:
            grupo.archivos_muy_grandes.append(ruta_completa)
            continue

        if len(grupo.archivos_validos) >= MAX_ARCHIVOS_POR_ID:
            grupo.archivos_excedente.append(ruta_completa)
            continue

        grupo.archivos_validos.append(ruta_completa)

    return grupos, no_reconocidos


def listar_meses_disponibles():
    if not os.path.isdir(CARPETA_FOTOS_BASE):
        return []
    carpetas = [
        nombre for nombre in sorted(os.listdir(CARPETA_FOTOS_BASE))
        if os.path.isdir(os.path.join(CARPETA_FOTOS_BASE, nombre))
    ]
    return carpetas


# ---------------------------------------------------------------------------
# PASO 2: REGISTRO PERSISTENTE (procesados.csv)
# ---------------------------------------------------------------------------


def cargar_registro() -> set[str]:
    procesados = set()
    if os.path.exists(ARCHIVO_REGISTRO):
        with open(ARCHIVO_REGISTRO, newline="", encoding="utf-8-sig") as f:
            lector = csv.DictReader(f)
            for fila in lector:
                if fila.get("estado") == "OK":
                    clave = f"{fila['id_cliente']}|{os.path.basename(fila['archivo'])}"
                    procesados.add(clave)
    return procesados


def registrar_resultado(id_cliente: str, archivo: str, estado: str, detalle: str = ""):
    existe = os.path.exists(ARCHIVO_REGISTRO)
    with open(ARCHIVO_REGISTRO, "a", newline="", encoding="utf-8-sig") as f:
        escritor = csv.writer(f)
        if not existe:
            escritor.writerow(["fecha", "id_cliente", "archivo", "estado", "detalle"])
        escritor.writerow([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            id_cliente,
            archivo,
            estado,
            detalle,
        ])


# ---------------------------------------------------------------------------
# PASO 3: AUTOMATIZACIÓN DE LA VENTANA (pywinauto + win32gui)
# ---------------------------------------------------------------------------


class AutomatizadorApp:
    def __init__(self):
        self.app = None
        self.ventana_principal = None
        self.ventana_principal_uia = None
        self.main_handle = None

    def conectar(self):
        import win32gui
        from pywinauto import Desktop, findwindows
        from pywinauto.application import Application
        from pywinauto.timings import wait_until_passes

        def _buscar_ventana_valida():
            elementos = findwindows.find_elements(
                title_re=TITULO_VENTANA_PRINCIPAL_REGEX, backend="win32"
            )
            validos = [e for e in elementos if win32gui.IsWindow(e.handle)]
            if not validos:
                raise findwindows.ElementNotFoundError(
                    "No se encontro ninguna ventana valida (no-fantasma)."
                )
            return validos[0]

        elemento_valido = wait_until_passes(
            TIMEOUT_ESPERA, 0.5, _buscar_ventana_valida
        )

        self.main_handle = elemento_valido.handle

        self.app = Application(backend="win32").connect(
            handle=self.main_handle, timeout=TIMEOUT_ESPERA
        )
        self.ventana_principal = self.app.window(handle=self.main_handle)
        self.ventana_principal.wait("exists ready", timeout=TIMEOUT_ESPERA)

        self.ventana_principal_uia = Desktop(backend="uia").window(
            handle=self.main_handle
        )
        self.ventana_principal_uia.wait("exists ready", timeout=TIMEOUT_ESPERA)

    def _obtener_ventana_top_valida(self):
        import win32gui
        ultimo_error = None
        for _ in range(20):
            try:
                ventana = self.app.top_window()
                if win32gui.IsWindow(ventana.handle):
                    return ventana
            except Exception as e:
                ultimo_error = e
            time.sleep(0.1)
        raise RuntimeError(f"No se pudo obtener una ventana activa valida. Error: {ultimo_error}")

    def _campo_edit_por_etiqueta(self, texto_etiqueta: str):
        todos = self.ventana_principal.descendants()

        etiqueta = None
        for ctrl in todos:
            try:
                if (ctrl.class_name().startswith("WindowsForms10.STATIC.")
                        and ctrl.window_text().strip() == texto_etiqueta):
                    etiqueta = ctrl
                    break
            except Exception:
                continue
        if etiqueta is None:
            raise RuntimeError(f"No se encontro la etiqueta {texto_etiqueta!r}")

        rect_etq = etiqueta.rectangle()
        centro_y_etq = (rect_etq.top + rect_etq.bottom) / 2

        mejor = None
        mejor_dist = None
        for ctrl in todos:
            try:
                if not ctrl.class_name().startswith("WindowsForms10.EDIT."):
                    continue
                r = ctrl.rectangle()
            except Exception:
                continue
            centro_y = (r.top + r.bottom) / 2
            if abs(centro_y - centro_y_etq) <= 10 and r.left >= rect_etq.left - 5:
                dist = r.left - rect_etq.right
                if mejor_dist is None or dist < mejor_dist:
                    mejor = ctrl
                    mejor_dist = dist

        if mejor is None:
            raise RuntimeError(f"Se encontro la etiqueta {texto_etiqueta!r} pero ningun Edit al lado.")
        return mejor

    def buscar_cliente(self, id_cliente: str) -> bool:
        try:
            self.ventana_principal.set_focus()
            time.sleep(PAUSA_CORTA)
        except Exception:
            pass

        campo_nro = self._campo_edit_por_etiqueta(NOMBRE_CAMPO_NRO)
        campo_nro.set_focus()
        campo_nro.set_edit_text("")
        time.sleep(PAUSA_CORTA)
        campo_nro.type_keys(id_cliente, with_spaces=True)
        campo_nro.type_keys("{ENTER}")
        time.sleep(PAUSA_CORTA * 2)

        texto_actual = campo_nro.window_text().strip()
        return texto_actual != ""

    def _buscar_boton(self, ventana, texto_boton: str):
        candidatos = ventana.descendants()
        for c in candidatos:
            try:
                clase = c.class_name()
                es_boton_valido = (
                    clase.startswith("WindowsForms10.BUTTON.") or clase == "Button"
                )
                if es_boton_valido and c.window_text().strip() == texto_boton:
                    return c
            except Exception:
                continue
        raise RuntimeError(f"No se encontro boton {texto_boton!r}")

    def _buscar_boton_toolstrip(self, texto_boton: str):
        for ctrl in self.ventana_principal_uia.descendants(control_type="Button"):
            try:
                if ctrl.window_text().strip() == texto_boton:
                    return ctrl
            except Exception:
                continue
        raise RuntimeError(f"No se encontro toolstrip button {texto_boton!r}")

    def _click_boton_toolstrip(self, texto_boton: str):
        handle_antes = self.app.top_window().handle

        for _ in range(2):
            self.ventana_principal.set_focus()
            time.sleep(PAUSA_CORTA)
            boton = self._buscar_boton_toolstrip(texto_boton)
            boton.click_input()
            time.sleep(PAUSA_CORTA * 2)

            try:
                if self.app.top_window().handle != handle_antes:
                    return
            except Exception:
                return

        raise RuntimeError(f"Se clickeo {texto_boton!r} pero no abrio nada nuevo.")

    def contar_imagenes_adjuntas(self, id_cliente: str = "") -> int:
        """Abre 'Ver imagen(es)':
        - Si sale la ventana 'No existen archivos adjuntos': la cierra y devuelve 0.
        - Si sale la ventana 'Imágenes del cliente' (tiene imágenes):
          la cierra de forma segura mediante WM_CLOSE dirigiéndose exclusivamente a ella
          (sin riesgo de cerrar la ventana principal) y devuelve 1.
        """
        import win32con
        import win32gui

        self._click_boton_toolstrip(NOMBRE_BOTON_VER_IMAGENES)
        time.sleep(PAUSA_CORTA * 2)

        ventana_emergente = None
        for _ in range(15):
            try:
                top = self._obtener_ventana_top_valida()
                if top.handle != self.main_handle:
                    ventana_emergente = top
                    break
            except Exception:
                pass
            time.sleep(0.2)

        if ventana_emergente is None:
            return 0

        # CASO 1: Ventana de "No existen archivos adjuntos" (MessageBox)
        try:
            boton_aceptar = self._buscar_boton(ventana_emergente, NOMBRE_BOTON_ACEPTAR_POPUP)
        except RuntimeError:
            boton_aceptar = None

        if boton_aceptar is not None:
            try:
                ventana_emergente.set_focus()
                time.sleep(0.2)
                ventana_emergente.type_keys("{ENTER}")
                time.sleep(PAUSA_CORTA)
            except Exception:
                pass

            if ventana_emergente.exists() and ventana_emergente.handle != self.main_handle:
                try:
                    boton_aceptar.click()
                    time.sleep(PAUSA_CORTA)
                except Exception:
                    pass

            return 0

        # CASO 2: Apareció la ventana "Imágenes del cliente" (SÍ tiene imágenes)
        try:
            h_popup = ventana_emergente.handle
            if h_popup != self.main_handle:
                win32gui.PostMessage(h_popup, win32con.WM_CLOSE, 0, 0)
                time.sleep(PAUSA_CORTA * 2)

            if ventana_emergente.exists() and ventana_emergente.handle != self.main_handle:
                ventana_emergente.set_focus()
                time.sleep(0.2)
                ventana_emergente.type_keys("%{F4}")
                time.sleep(PAUSA_CORTA)
        except Exception as e:
            print(f"Aviso al cerrar ventana emergente: {e}")

        try:
            self.ventana_principal.set_focus()
        except Exception:
            pass

        return 1

    def _click_boton_explorar(self, ventana_adjuntar):
        for _ in range(2):
            ventana_adjuntar.set_focus()
            time.sleep(PAUSA_CORTA)
            boton_explorar = self._buscar_boton(ventana_adjuntar, NOMBRE_BOTON_EXPLORAR)
            boton_explorar.click_input()
            time.sleep(PAUSA_CORTA * 2)

            try:
                self.app.window(class_name="#32770").wait("exists", timeout=3)
                return
            except Exception:
                continue

        raise RuntimeError("No aparecio el dialogo del explorador.")

    def _escribir_ruta_en_dialogo(self, dialogo, ruta_absoluta: str):
        """Escribe la ruta del archivo en el campo 'Nombre:' del dialogo del
        Explorador de Windows.

        En vez de simular Alt+N (que depende de donde haya quedado el foco
        la ultima vez que se abrio el dialogo en ese equipo -a veces queda
        en la barra de direcciones/breadcrumb de arriba y todo lo tecleado
        se va ahi en lugar del campo de nombre-), se ubica directamente el
        control real por su AutomationId. El campo 'Nombre:' del dialogo
        moderno de Windows (comdlg32 / IFileDialog) tiene AutomationId
        '1001' de forma fija, sin importar el idioma de Windows.
        """
        from pywinauto import Desktop

        try:
            dialogo_uia = Desktop(backend="uia").window(handle=dialogo.handle)
            dialogo_uia.wait("exists ready", timeout=TIMEOUT_ESPERA)

            campo_nombre = dialogo_uia.child_window(auto_id="1001", control_type="Edit")
            campo_nombre.wait("exists ready", timeout=5)

            campo_nombre.set_focus()
            time.sleep(PAUSA_CORTA)
            campo_nombre.set_edit_text(ruta_absoluta)
            time.sleep(PAUSA_CORTA)
            campo_nombre.type_keys("{ENTER}")
            return
        except Exception:
            pass

        # --- Fallback: metodo antiguo por si en algun equipo no aparece
        # el control con ese AutomationId (version de Windows distinta, etc).
        dialogo.set_focus()
        time.sleep(PAUSA_CORTA)
        dialogo.type_keys("%n", set_foreground=True)
        time.sleep(PAUSA_CORTA)
        dialogo.type_keys("^a{BACKSPACE}", set_foreground=True)
        time.sleep(PAUSA_CORTA)
        dialogo.type_keys(ruta_absoluta, with_spaces=True, set_foreground=True)
        time.sleep(PAUSA_CORTA)
        dialogo.type_keys("{ENTER}")

    def adjuntar_archivo(self, ruta_archivo: str, carpeta_mes: str = ""):
        """Abre 'Adjuntar imágen', navega mediante explorador de Windows,
        forzando foco en la casilla 'Nombre:' con Alt+N y limpiando el campo con Ctrl+A + Backspace."""
        self._click_boton_toolstrip(NOMBRE_BOTON_ADJUNTAR)
        time.sleep(PAUSA_CORTA * 2)

        ventana_adjuntar = self._obtener_ventana_top_valida()
        ventana_adjuntar.wait("exists ready", timeout=TIMEOUT_ESPERA)

        self._click_boton_explorar(ventana_adjuntar)

        dialogo = self.app.window(class_name="#32770")
        dialogo.wait("exists ready", timeout=TIMEOUT_ESPERA)

        ruta_absoluta = os.path.abspath(ruta_archivo)

        dialogo.set_focus()
        time.sleep(PAUSA_CORTA)

        self._escribir_ruta_en_dialogo(dialogo, ruta_absoluta)
        time.sleep(PAUSA_CORTA * 2)

        # 4. Presiona el botón "Agregar"
        boton_agregar = None
        for _ in range(10):
            try:
                boton_agregar = self._buscar_boton(ventana_adjuntar, NOMBRE_BOTON_AGREGAR)
                break
            except RuntimeError:
                time.sleep(PAUSA_CORTA)

        if boton_agregar is None:
            raise RuntimeError("No aparecio el boton 'Agregar'.")

        boton_agregar.click_input()
        time.sleep(PAUSA_CORTA * 2)

        # 5. Cierra la ventana emergente de adjuntar
        boton_cerrar = self._buscar_boton(ventana_adjuntar, NOMBRE_BOTON_CERRAR_POPUP)
        boton_cerrar.click()
        time.sleep(PAUSA_CORTA)


# ---------------------------------------------------------------------------
# PASO 4: FLUJO PRINCIPAL Y REPORTE EXCEL
# ---------------------------------------------------------------------------


def procesar_mes(nombre_mes: str, modo_solo_revisar: bool):
    carpeta_mes = os.path.join(CARPETA_FOTOS_BASE, nombre_mes)
    grupos, no_reconocidos = escanear_carpeta(carpeta_mes)
    ya_procesados = cargar_registro()

    resultados: list[ResultadoID] = []

    if not modo_solo_revisar:
        automatizador = AutomatizadorApp()
        print("Conectando con la aplicacion...")
        automatizador.conectar()
    else:
        automatizador = None

    total = len(grupos)
    for i, (id_cliente, grupo) in enumerate(grupos.items(), start=1):
        print(f"[{i}/{total}] Procesando ID {id_cliente} ...")

        if grupo.archivos_excedente:
            for archivo in grupo.archivos_excedente:
                registrar_resultado(id_cliente, archivo, "EXCEDENTE",
                                     f"Mas de {MAX_ARCHIVOS_POR_ID} archivos para este ID")
            resultados.append(ResultadoID(id_cliente, "EXCEDENTE",
                               f"Excedente: {len(grupo.archivos_excedente)} archivo(s) ignorado(s)"))

        if grupo.archivos_muy_grandes:
            for archivo in grupo.archivos_muy_grandes:
                registrar_resultado(id_cliente, archivo, "MUY_GRANDE",
                                     f"Supera los {MAX_TAMANO_MB}MB")
            resultados.append(ResultadoID(id_cliente, "MUY_GRANDE",
                               f"{len(grupo.archivos_muy_grandes)} archivo(s) > {MAX_TAMANO_MB}MB"))

        archivos_pendientes = [
            a for a in grupo.archivos_validos
            if f"{id_cliente}|{os.path.basename(a)}" not in ya_procesados
        ]

        if not archivos_pendientes:
            continue

        if modo_solo_revisar:
            resultados.append(ResultadoID(id_cliente, "PENDIENTE",
                               f"{len(archivos_pendientes)} archivo(s) pendientes",
                               archivos_pendientes))
            continue

        try:
            encontrado = automatizador.buscar_cliente(id_cliente)
            if not encontrado:
                for a in archivos_pendientes:
                    registrar_resultado(id_cliente, a, "NO_ENCONTRADO", "")
                resultados.append(ResultadoID(id_cliente, "NO_ENCONTRADO", "ID no encontrado en el sistema"))
                continue

            filas_antes = automatizador.contar_imagenes_adjuntas(id_cliente)
            if filas_antes > 0:
                for a in archivos_pendientes:
                    registrar_resultado(id_cliente, a, "OMITIDO_YA_TENIA",
                                         "El cliente ya tenia imagenes adjuntas")
                resultados.append(ResultadoID(id_cliente, "OMITIDO_YA_TENIA",
                                   "El cliente ya tenia imagenes en el sistema"))
                continue

            for archivo in archivos_pendientes:
                automatizador.adjuntar_archivo(archivo, carpeta_mes)

            for a in archivos_pendientes:
                registrar_resultado(id_cliente, a, "OK", "")
            resultados.append(ResultadoID(id_cliente, "OK",
                               f"Subida exitosa de {len(archivos_pendientes)} archivo(s)"))

        except Exception as e:
            traceback.print_exc()
            for a in archivos_pendientes:
                registrar_resultado(id_cliente, a, "ERROR", str(e))
            resultados.append(ResultadoID(id_cliente, "ERROR", str(e)))

    archivo_excel = generar_reporte_excel(nombre_mes, resultados, no_reconocidos)
    return resultados, no_reconocidos, archivo_excel


def generar_reporte_excel(nombre_mes: str, resultados: list[ResultadoID], no_reconocidos: list[str]) -> str:
    nombre_limpio = re.sub(r'[^\w\-_\. ]', '_', nombre_mes)
    nombre_archivo = f"reporte_{nombre_limpio}.xlsx"

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Carga"

        headers = ["ID", "¿Se subió?", "¿Necesita revisión?", "Detalle"]
        ws.append(headers)

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r in resultados:
            if r.estado == "OK":
                se_subio = "Sí"
                revision = "No"
            elif r.estado == "EXCEDENTE":
                se_subio = "Sí"
                revision = "Sí"
            else:
                se_subio = "No"
                revision = "Sí"

            ws.append([r.id_cliente, se_subio, revision, r.detalle])

        for archivo in no_reconocidos:
            ws.append([archivo, "No", "Sí", "Nombre de archivo no reconocido (formato no numérico)"])

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
                if cell.row > 1 and col_letter in ("A", "B", "C"):
                    cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        wb.save(nombre_archivo)
        return nombre_archivo

    except ImportError:
        nombre_csv = f"reporte_{nombre_limpio}.csv"
        with open(nombre_csv, "w", newline="", encoding="utf-8-sig") as f:
            escritor = csv.writer(f)
            escritor.writerow(["ID", "¿Se subió?", "¿Necesita revisión?", "Detalle"])
            for r in resultados:
                se_subio = "Sí" if r.estado in ("OK", "EXCEDENTE") else "No"
                revision = "No" if r.estado == "OK" else "Sí"
                escritor.writerow([r.id_cliente, se_subio, revision, r.detalle])
            for archivo in no_reconocidos:
                escritor.writerow([archivo, "No", "Sí", "Nombre de archivo no reconocido"])
        return nombre_csv


# ---------------------------------------------------------------------------
# PASO 5: MENÚ DE CONSOLA
# ---------------------------------------------------------------------------


def elegir_mes() -> str | None:
    meses = listar_meses_disponibles()
    if not meses:
        print(f"No se encontraron carpetas de mes dentro de:\n  {CARPETA_FOTOS_BASE}")
        return None

    print("\nMeses disponibles:")
    for i, mes in enumerate(meses, start=1):
        print(f"  {i}. {mes}")

    seleccion = input("\nElige el numero del mes: ").strip()
    if not seleccion.isdigit() or not (1 <= int(seleccion) <= len(meses)):
        print("Opcion invalida.")
        return None
    return meses[int(seleccion) - 1]


def menu_principal():
    while True:
        print("\n" + "=" * 50)
        print("  CARGA DE FOTOS - REFACTURACIONES")
        print("=" * 50)
        print("  1. Subir imagenes de un mes")
        print("  2. Revisar estado (sin subir nada)")
        print("  3. Salir")
        opcion = input("\nElige una opcion: ").strip()

        if opcion == "1":
            mes = elegir_mes()
            if mes:
                print(f"\nProcesando '{mes}'... esto puede tomar varios minutos.\n")
                resultados, no_reconocidos, archivo_reporte = procesar_mes(mes, modo_solo_revisar=False)
                mostrar_resumen(resultados, no_reconocidos, archivo_reporte)

        elif opcion == "2":
            mes = elegir_mes()
            if mes:
                print(f"\nRevisando '{mes}'...\n")
                resultados, no_reconocidos, archivo_reporte = procesar_mes(mes, modo_solo_revisar=True)
                mostrar_resumen(resultados, no_reconocidos, archivo_reporte)

        elif opcion == "3":
            print("Saliendo...")
            break
        else:
            print("Opcion invalida, intenta de nuevo.")

        input("\nPresiona ENTER para volver al menu...")


def mostrar_resumen(resultados, no_reconocidos, archivo_reporte):
    conteos = {}
    for r in resultados:
        conteos[r.estado] = conteos.get(r.estado, 0) + 1
    print("\n--- RESUMEN ---")
    if not resultados:
        print("  No habia nada nuevo que procesar.")
    for estado, cantidad in conteos.items():
        print(f"  {estado}: {cantidad}")
    if no_reconocidos:
        print(f"  Archivos con nombre no reconocido: {len(no_reconocidos)}")
    print(f"\nReporte Excel guardado en: {archivo_reporte}")


if __name__ == "__main__":
    try:
        menu_principal()
    except Exception:
        print("\n\nOcurrio un error inesperado:")
        traceback.print_exc()
        input("\nPresiona ENTER para cerrar...")
        sys.exit(1)